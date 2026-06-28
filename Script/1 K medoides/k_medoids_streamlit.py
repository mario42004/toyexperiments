# streamlit_app.py
# App Streamlit: input = texto, output = archivo madre Excel etiquetado.
from __future__ import annotations

import io
import json
import re
import urllib.error
import urllib.request
from typing import Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from spanish_kmedoids_10 import (
    ICOM_KEYWORDS,
    LEY_19_2013_KEYWORDS,
    cosine_distance_matrix,
    improve_medoids,
    initialize_medoids,
    keyword_labels,
    medoid_cost,
    normalize_for_keywords,
    tfidf_vectors,
)


# --------------------------
# Utilidades
# --------------------------
STOPWORDS = {
    "para", "por", "con", "sin", "del", "las", "los", "una", "uno", "unos",
    "unas", "que", "como", "sobre", "entre", "desde", "hasta", "esta", "este",
    "estos", "estas", "tambien", "donde", "cuando", "cada", "otros", "otras",
    "informacion", "transparencia", "dimension", "codigo", "marco", "temas",
    "institucion", "institucional", "museo", "museos",
}


def is_legible_paragraph(text: str) -> bool:
    letters = re.findall(r"[^\W\d_]", text)
    visible_chars = re.findall(r"\S", text)
    words = re.findall(r"[^\W\d_]{3,}", text.lower())
    short_tokens = re.findall(r"\b[^\W\d_]{1,2}\b", text.lower())
    numeric_tokens = re.findall(r"\b\d+(?:[.,]\d+)?\b", text)
    table_tokens = re.findall(r"\b(?:td|tr|lt|gt)\b", text.lower())
    all_tokens = re.findall(r"\b\w+\b", text.lower())

    if len(letters) < 12:
        return False
    if len(words) < 4:
        return False
    if len(table_tokens) >= 3:
        return False
    if len(numeric_tokens) > len(words):
        return False
    if len(short_tokens) / max(len(all_tokens), 1) > 0.35:
        return False
    return len(letters) / max(len(visible_chars), 1) >= 0.45


def split_sentences_or_paragraphs(text: str) -> List[str]:
    """
    Divide el texto en oraciones o parrafos.
    Usa saltos de linea si existen; si no, divide por signos de puntuacion.
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) > 1:
        return [line for line in lines if is_legible_paragraph(line)]

    sentences = [s.strip() for s in re.split(r"[.!?]\s+", text) if s.strip()]
    return [sentence for sentence in sentences if is_legible_paragraph(sentence)]


def elbow_candidate_values(total_items: int, min_k: int = 30, max_k: int = 90) -> List[int]:
    if total_items <= 0:
        return []
    if total_items <= min_k:
        return [total_items]

    upper = min(max_k, total_items)
    candidates = list(range(min_k, upper + 1, 10))
    if upper not in candidates:
        candidates.append(upper)
    return sorted(set(candidates))


def choose_elbow_k(cost_rows: List[Dict[str, object]]) -> int:
    if len(cost_rows) <= 2:
        return int(cost_rows[0]["k"])

    first = cost_rows[0]
    last = cost_rows[-1]
    x1, y1 = float(first["k"]), float(first["coste"])
    x2, y2 = float(last["k"]), float(last["coste"])
    denominator = ((y2 - y1) ** 2 + (x2 - x1) ** 2) ** 0.5
    if denominator == 0:
        return int(first["k"])

    best_row = first
    best_distance = -1.0
    for row in cost_rows[1:-1]:
        x0, y0 = float(row["k"]), float(row["coste"])
        distance = abs((y2 - y1) * x0 - (x2 - x1) * y0 + x2 * y1 - y2 * x1) / denominator
        if distance > best_distance:
            best_distance = distance
            best_row = row
    return int(best_row["k"])


def select_k_medoids_with_elbow(
    paragraphs: List[str],
    min_k: int = 30,
    max_k: int = 90,
) -> Tuple[List[str], List[Dict[str, object]], Dict[str, object]]:
    total_items = len(paragraphs)
    candidates = elbow_candidate_values(total_items, min_k=min_k, max_k=max_k)
    if not candidates:
        return [], [], {
            "parrafos_validos": 0,
            "k_minimo": min_k,
            "k_maximo": max_k,
            "k_optimo": 0,
            "criterio": "sin parrafos validos",
        }

    if total_items <= min_k:
        selected_k = total_items
        return paragraphs, [
            {
                "k": selected_k,
                "coste": 0.0,
                "mejora_absoluta": 0.0,
                "mejora_relativa": 0.0,
                "seleccionado": True,
            }
        ], {
            "parrafos_validos": total_items,
            "k_minimo": min_k,
            "k_maximo": max_k,
            "k_optimo": selected_k,
            "criterio": "se devuelven todos los parrafos validos porque N es menor o igual al minimo",
        }

    vectors = tfidf_vectors(paragraphs)
    distances = cosine_distance_matrix(vectors)
    cost_rows = []
    medoids_by_k = {}

    for candidate_k in candidates:
        medoids = initialize_medoids(distances, candidate_k)
        medoids = improve_medoids(distances, medoids, max_iter=5)
        cost = medoid_cost(distances, medoids)
        medoids_by_k[candidate_k] = medoids
        cost_rows.append(
            {
                "k": candidate_k,
                "coste": round(cost, 6),
                "mejora_absoluta": 0.0,
                "mejora_relativa": 0.0,
                "seleccionado": False,
            }
        )

    previous_cost = None
    for row in cost_rows:
        current_cost = float(row["coste"])
        if previous_cost is not None:
            improvement = previous_cost - current_cost
            row["mejora_absoluta"] = round(improvement, 6)
            row["mejora_relativa"] = round(improvement / previous_cost, 6) if previous_cost else 0.0
        previous_cost = current_cost

    selected_k = choose_elbow_k(cost_rows)
    for row in cost_rows:
        row["seleccionado"] = int(row["k"]) == selected_k

    selected_items = [paragraphs[index] for index in medoids_by_k[selected_k]]
    return selected_items, cost_rows, {
        "parrafos_validos": total_items,
        "k_minimo": min_k,
        "k_maximo": max_k,
        "k_optimo": selected_k,
        "criterio": "metodo del codo sobre coste intra-cluster de K-medoides",
    }


def make_excel_bytes(
    rows: List[List],
    header: List[str],
    elbow_rows: Optional[List[Dict[str, object]]] = None,
    elbow_summary: Optional[Dict[str, object]] = None,
) -> bytes:
    buf = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "archivo_madre"
    ws.append(header)

    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [90, 35, 35, 35, 35, 55, 55, 65, 65, 35]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    if elbow_rows is not None and elbow_summary is not None:
        ws_elbow = wb.create_sheet("metodo_codo")
        ws_elbow.append(["criterio", "valor"])
        for key, value in elbow_summary.items():
            ws_elbow.append([key, value])
        ws_elbow.append([])
        ws_elbow.append(["k", "coste", "mejora_absoluta", "mejora_relativa", "seleccionado"])
        for row in elbow_rows:
            ws_elbow.append(
                [
                    row["k"],
                    row["coste"],
                    row["mejora_absoluta"],
                    row["mejora_relativa"],
                    "si" if row["seleccionado"] else "",
                ]
            )
        for row in ws_elbow.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for cell in ws_elbow[1]:
            cell.font = Font(bold=True)
        ws_elbow.column_dimensions["A"].width = 28
        ws_elbow.column_dimensions["B"].width = 45
        ws_elbow.column_dimensions["C"].width = 20
        ws_elbow.column_dimensions["D"].width = 20
        ws_elbow.column_dimensions["E"].width = 15

    wb.save(buf)
    return buf.getvalue()


def label_to_display(label: str) -> str:
    icom_labels = {
        "dimension_01_transparencia_institucional": "transparencia institucional",
        "dimension_02_transparencia_economica": "transparencia economica",
        "dimension_03_transparencia_colecciones": "transparencia sobre las colecciones",
        "dimension_04_transparencia_procedencia_legalidad": "transparencia sobre procedencia y legalidad",
        "dimension_05_transparencia_cientifica": "transparencia cientifica",
        "dimension_06_transparencia_educativa": "transparencia educativa",
        "dimension_07_transparencia_acceso": "transparencia sobre el acceso",
        "dimension_08_transparencia_servicio_publico": "transparencia del servicio publico",
        "dimension_09_transparencia_social": "transparencia social",
        "dimension_10_transparencia_cooperacion_institucional": "transparencia en cooperacion institucional",
        "dimension_11_transparencia_juridica": "transparencia juridica",
        "dimension_12_transparencia_etica": "transparencia etica",
    }
    if label in icom_labels:
        return icom_labels[label]
    label = re.sub(r"^dimension_\d+_", "", label)
    prefixes = [
        ("informacion_institucional_", "informacion institucional"),
        ("informacion_organizativa_", "informacion organizativa"),
        ("planificacion_", "informacion sobre planificacion"),
        ("evaluacion_", "evaluacion del grado de cumplimiento"),
        ("normativa_", "normativa que las rige"),
        ("funciones_", "funciones que desarrolla"),
        ("informacion_juridica_", "informacion juridica relevante"),
        ("contratos_", "contratos"),
        ("convenios_", "convenios"),
        ("subvenciones_", "subvenciones y ayudas publicas"),
        ("presupuestos_", "presupuestos"),
        ("cuentas_anuales_", "cuentas anuales"),
        ("estadisticas_", "estadisticas"),
        ("acceso_", "acceso a la informacion"),
        ("buen_gobierno_", "buen gobierno"),
    ]
    for prefix, group in prefixes:
        if label.startswith(prefix):
            topic = label.removeprefix(prefix).replace("_", " ")
            return f"{group}, {topic}"
    return label.replace("_", " ")


def join_display_labels(labels: List[str]) -> str:
    return ", ".join(label_to_display(label) for label in labels)


def display_taxonomy(taxonomy: Dict[str, List[str]]) -> Dict[str, List[str]]:
    return {label_to_display(label): keywords for label, keywords in taxonomy.items()}


def tokens_for_candidate_selection(text: str) -> set[str]:
    normalized = normalize_for_keywords(text.replace("_", " "))
    tokens = re.findall(r"\b[a-z0-9]{3,}\b", normalized)
    return {token for token in tokens if token not in STOPWORDS}


def select_candidate_taxonomies(
    paragraph: str,
    law_taxonomy: Dict[str, List[str]],
    icom_taxonomy: Dict[str, List[str]],
    max_candidates: int = 18,
) -> tuple[Dict[str, List[str]], Dict[str, List[str]]]:
    paragraph_tokens = tokens_for_candidate_selection(paragraph)
    scored = []
    for marco, taxonomy in [("ley", law_taxonomy), ("icom", icom_taxonomy)]:
        for label, keywords in taxonomy.items():
            display_label = label_to_display(label)
            candidate_tokens = tokens_for_candidate_selection(" ".join([display_label, *keywords]))
            overlap = paragraph_tokens.intersection(candidate_tokens)
            if overlap:
                score = len(overlap)
                scored.append((score, marco, display_label, keywords))

    if not scored:
        return {}, display_taxonomy(icom_taxonomy)

    scored.sort(reverse=True)
    law_candidates = {}
    icom_candidates = {}
    for _, marco, label, keywords in scored[:max_candidates]:
        if marco == "ley":
            law_candidates[label] = keywords
        else:
            icom_candidates[label] = keywords
    return law_candidates, icom_candidates


def taxonomy_to_text(taxonomy: Dict[str, List[str]]) -> str:
    return "\n".join(
        f"{label_to_display(label)} = {', '.join(keywords)}"
        for label, keywords in taxonomy.items()
    )


def text_to_taxonomy(text: str) -> Dict[str, List[str]]:
    taxonomy = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue

        if "=" in line:
            label, raw_keywords = line.split("=", 1)
        elif ":" in line:
            label, raw_keywords = line.split(":", 1)
        else:
            continue

        label = label.strip()
        raw_keywords = raw_keywords.strip()
        if not label or not raw_keywords:
            continue

        keywords = [kw.strip() for kw in raw_keywords.split(",") if kw.strip()]
        if keywords:
            taxonomy[label] = keywords

    return taxonomy


def taxonomy_options_text(
    law_taxonomy: Dict[str, List[str]],
    icom_taxonomy: Dict[str, List[str]],
) -> str:
    lines = ["LEY_19_2013:"]
    for label, keywords in law_taxonomy.items():
        lines.append(f"- {label}: {', '.join(keywords)}")
    lines.append("CODIGO_MUSEOS:")
    for label, keywords in icom_taxonomy.items():
        lines.append(f"- {label}: {', '.join(keywords)}")
    return "\n".join(lines)


def keyword_hits_for_label(text: str, taxonomy: Dict[str, List[str]], label: str) -> List[str]:
    if label == "indeterminado" or label not in taxonomy:
        return []
    normalized_text = normalize_for_keywords(text)
    return [
        keyword
        for keyword in taxonomy[label]
        if normalize_for_keywords(keyword) in normalized_text
    ]


def textual_ia_justification(
    framework_name: str,
    label: str,
    hits: List[str],
    total_keywords: int,
) -> str:
    if label == "indeterminado":
        return (
            f"{framework_name}: indeterminado. "
            "No se identificaron indicios textuales suficientes para asignar una etiqueta cerrada."
        )
    if not hits:
        return (
            f"{framework_name}: {label}. "
            "No se detectaron indicios lexicos literales de la lista; la asignacion queda sustentada por correspondencia semantica del parrafo."
        )
    return (
        f"{framework_name}: {label}. "
        f"Indicios detectados {len(hits)}/{total_keywords}: {', '.join(hits)}."
    )


def fallback_reasoning(framework_name: str, label: str) -> str:
    if label == "indeterminado":
        return (
            f"No se asigna etiqueta para {framework_name} porque el parrafo no presenta una relacion normativa suficientemente clara con las categorias disponibles."
        )
    return (
        f"Se asigna {label} para {framework_name} porque el contenido del parrafo guarda correspondencia tematica con esa categoria normativa."
    )


def contains_english_markers(text: str) -> bool:
    english_markers = {
        "because", "paragraph", "label", "category", "evidence", "text", "matches",
        "does not", "information", "public", "transparency", "collection", "reason",
        "assigned", "according", "law", "code", "museum", "museums",
    }
    normalized = normalize_for_keywords(text)
    words = set(re.findall(r"\b[a-z]{3,}\b", normalized))
    return bool(words.intersection(english_markers))


def spanish_reasoning_or_fallback(text: str, framework_name: str, label: str) -> str:
    text = text.strip()
    if not text or contains_english_markers(text):
        return fallback_reasoning(framework_name, label)
    return text


def extract_json_object(text: str) -> Dict[str, object]:
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if not match:
            raise
        return json.loads(match.group(0))


def classify_with_ollama(
    paragraph: str,
    law_taxonomy: Dict[str, List[str]],
    icom_taxonomy: Dict[str, List[str]],
    model: str,
    ollama_url: str,
) -> Dict[str, str]:
    law_display_taxonomy = display_taxonomy(law_taxonomy)
    icom_display_taxonomy = display_taxonomy(icom_taxonomy)
    law_prompt_taxonomy, icom_prompt_taxonomy = select_candidate_taxonomies(
        paragraph,
        law_taxonomy,
        icom_taxonomy,
    )
    valid_law_labels = set(law_display_taxonomy)
    valid_icom_labels = set(icom_display_taxonomy)
    prompt = f"""
Eres un clasificador academico para una tesis doctoral.
Debes leer el parrafo y asignar DOS etiquetas independientes dentro de taxonomias cerradas:
1. Una etiqueta segun Ley 19/2013.
2. Una etiqueta segun Codigo deontologico de museos.

No inventes etiquetas. Si ninguna etiqueta corresponde razonablemente en una taxonomia, usa "indeterminado" para esa taxonomia.
Actua como una persona codificadora que revisa evidencias textuales y asigna categorias normativas.
Todos los campos textuales deben estar escritos exclusivamente en espanol.
No uses palabras, frases ni conectores en ingles.
No traduzcas los nombres de las etiquetas: mantenlos exactamente como aparecen en la taxonomia.

Taxonomia cerrada:
{taxonomy_options_text(law_prompt_taxonomy, icom_prompt_taxonomy)}

Parrafo:
\"\"\"{paragraph}\"\"\"

Responde solo JSON valido con este esquema:
{{
  "etiqueta_ia_ley_19_2013": "una etiqueta exacta de LEY_19_2013 o indeterminado",
  "etiqueta_ia_codigo_deontologico": "una etiqueta exacta de CODIGO_MUSEOS o indeterminado",
  "razonamiento_ia_ley_19_2013": "una explicacion breve y clara de por que el parrafo encaja o no encaja con la etiqueta de Ley 19/2013",
  "razonamiento_ia_codigo_deontologico": "una explicacion breve y clara de por que el parrafo encaja o no encaja con la etiqueta del Codigo deontologico"
}}
"""
    payload = {
        "model": model,
        "prompt": prompt,
        "stream": False,
        "format": "json",
        "options": {"temperature": 0.0},
    }
    request = urllib.request.Request(
        ollama_url.rstrip("/") + "/api/generate",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            raw = json.loads(response.read().decode("utf-8", errors="replace"))
        parsed = extract_json_object(raw.get("response", "{}"))
    except (urllib.error.URLError, TimeoutError, json.JSONDecodeError, KeyError) as exc:
        return {
            "etiqueta_ia_ley_19_2013": "indeterminado",
            "etiqueta_ia_codigo_deontologico": "indeterminado",
            "justificacion_ia_ley_19_2013": f"Error al consultar Ollama: {exc}",
            "justificacion_ia_codigo_deontologico": f"Error al consultar Ollama: {exc}",
            "razonamiento_ia_ley_19_2013": "No se pudo generar razonamiento porque fallo la consulta a Ollama.",
            "razonamiento_ia_codigo_deontologico": "No se pudo generar razonamiento porque fallo la consulta a Ollama.",
        }

    law_label = str(parsed.get("etiqueta_ia_ley_19_2013", "indeterminado")).strip()
    icom_label = str(parsed.get("etiqueta_ia_codigo_deontologico", "indeterminado")).strip()
    if law_label not in valid_law_labels:
        law_label = "indeterminado"
    if icom_label not in valid_icom_labels:
        icom_label = "indeterminado"

    law_hits = keyword_hits_for_label(paragraph, law_taxonomy, law_label)
    icom_hits = keyword_hits_for_label(paragraph, icom_taxonomy, icom_label)
    law_reasoning = str(parsed.get("razonamiento_ia_ley_19_2013", "")).strip()
    icom_reasoning = str(parsed.get("razonamiento_ia_codigo_deontologico", "")).strip()

    return {
        "etiqueta_ia_ley_19_2013": law_label,
        "etiqueta_ia_codigo_deontologico": icom_label,
        "justificacion_ia_ley_19_2013": textual_ia_justification(
            "Ley 19/2013",
            law_label,
            law_hits,
            len(law_taxonomy.get(law_label, [])),
        ),
        "justificacion_ia_codigo_deontologico": textual_ia_justification(
            "Codigo deontologico",
            icom_label,
            icom_hits,
            len(icom_taxonomy.get(icom_label, [])),
        ),
        "razonamiento_ia_ley_19_2013": spanish_reasoning_or_fallback(
            law_reasoning,
            "Ley 19/2013",
            law_label,
        ),
        "razonamiento_ia_codigo_deontologico": spanish_reasoning_or_fallback(
            icom_reasoning,
            "Codigo deontologico",
            icom_label,
        ),
    }


def automated_labeling_columns(
    text: str,
    law_taxonomy: Dict[str, List[str]],
    icom_taxonomy: Dict[str, List[str]],
    model: str,
    ollama_url: str,
) -> Dict[str, str]:
    law_labels = keyword_labels(text, law_taxonomy)
    icom_labels = keyword_labels(text, icom_taxonomy)
    ia_labels = classify_with_ollama(text, law_taxonomy, icom_taxonomy, model, ollama_url)
    final_candidates = [
        ia_labels["etiqueta_ia_ley_19_2013"],
        ia_labels["etiqueta_ia_codigo_deontologico"],
    ]
    final_label = ", ".join(label for label in final_candidates if label != "indeterminado")
    if not final_label:
        dict_labels = [label_to_display(label) for label in law_labels + icom_labels]
        final_label = ", ".join(dict_labels) if dict_labels else "indeterminado"

    return {
        "ley_diccionario": join_display_labels(law_labels),
        "codigo_diccionario": join_display_labels(icom_labels),
        **ia_labels,
        "etiqueta_final": final_label,
    }


# --------------------------
# UI
# --------------------------
st.set_page_config(
    page_title="Archivo madre etiquetado",
    page_icon=":material/table_chart:",
    layout="centered",
)

st.markdown(
    """
    <style>
    div.stButton > button[kind="primary"] {
        background-color: #16a34a;
        border: 1px solid #15803d;
        border-radius: 8px;
        color: white;
        font-size: 1.35rem;
        font-weight: 700;
        min-height: 4rem;
    }
    div.stButton > button[kind="primary"]:hover {
        background-color: #15803d;
        border-color: #166534;
        color: white;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Archivo madre etiquetado")
st.write(
    "Esta herramienta lee una memoria, separa sus parrafos y genera un solo "
    "Excel con las etiquetas normativas."
)

with st.expander("Que significa cada parte", expanded=True):
    st.markdown(
        """
        - **Pegar texto**: pega aqui el contenido de una memoria si quieres probar sin subir archivo.
        - **Subir archivo .txt/.md**: carga una memoria completa desde tu ordenador.
        - **Metodo del codo**: calcula automaticamente el numero de parrafos representativos, con minimo 30 y maximo 90.
        - **Nombre de la ley o marco 1**: cambia aqui Ley 19/2013 por cualquier otra ley o marco que quieras analizar.
        - **Nombre del marco 2**: puedes dejar Codigo deontologico ICOM o cambiarlo por otro marco.
        - **Etiquetas y palabras clave**: escribe una etiqueta por linea con este formato: etiqueta = palabra1, palabra2, palabra3.
        - **Etiqueta final**: resume las etiquetas IA asignadas para los dos marcos normativos.
        - **Restaurar marco 1 / Restaurar marco 2**: vuelve a poner las etiquetas originales si cambiaste algo y quieres empezar de nuevo.
        - **Procesar**: crea el Excel final.
        - **Archivo madre etiquetado**: el unico archivo de salida. Contiene una fila por parrafo, evidencia por diccionario, clasificacion IA por marco, indicios textuales, razonamiento IA y etiqueta final. Incluye una hoja aparte con el detalle del metodo del codo.
        """
    )

with st.expander("Nota metodologica", expanded=False):
    st.markdown(
        """
        La version actual usa una clasificacion automatizada con taxonomia
        cerrada: el diccionario conserva evidencia textual y la IA asigna
        una etiqueta interpretativa para Ley 19/2013 y otra para el Codigo
        deontologico, sin poder inventar categorias nuevas.
        """
    )

with st.container():
    st.subheader("Editar leyes y etiquetas")
    st.caption(
        "Formato: una etiqueta por linea. Ejemplo: "
        "presupuesto_aprobado = presupuesto, cuentas, gasto"
    )

    if "law_name" not in st.session_state:
        st.session_state.law_name = "ley 19/2013"
    if "icom_name" not in st.session_state:
        st.session_state.icom_name = "codigo deontologico icom"
    if "law_taxonomy_text" not in st.session_state:
        st.session_state.law_taxonomy_text = taxonomy_to_text(LEY_19_2013_KEYWORDS)
    if "icom_taxonomy_text" not in st.session_state:
        st.session_state.icom_taxonomy_text = taxonomy_to_text(ICOM_KEYWORDS)

    law_name = st.text_input(
        "Nombre de la ley o marco 1",
        value=st.session_state.law_name,
        help="Este nombre sera el titulo de la columna en el Excel. Puedes cambiarlo por otra ley.",
    )
    icom_name = st.text_input(
        "Nombre del marco 2",
        value=st.session_state.icom_name,
        help="Este nombre sera el titulo de la segunda columna de etiquetas.",
    )

    with st.expander(law_name, expanded=True):
        law_taxonomy_text = st.text_area(
            "Etiquetas y palabras clave del marco 1",
            value=st.session_state.law_taxonomy_text,
            height=320,
            help="Una etiqueta por linea. Ejemplo: contratos_publicos = contrato, licitacion, adjudicacion",
        )

    with st.expander(icom_name, expanded=True):
        icom_taxonomy_text = st.text_area(
            "Etiquetas y palabras clave del marco 2",
            value=st.session_state.icom_taxonomy_text,
            height=320,
            help="Una etiqueta por linea. Ejemplo: accesibilidad = accesible, discapacidad, inclusion",
        )

    col_reset_ley, col_reset_icom = st.columns(2)
    with col_reset_ley:
        if st.button("Restaurar marco 1"):
            st.session_state.law_name = "ley 19/2013"
            st.session_state.law_taxonomy_text = taxonomy_to_text(LEY_19_2013_KEYWORDS)
            st.rerun()
    with col_reset_icom:
        if st.button("Restaurar marco 2"):
            st.session_state.icom_name = "codigo deontologico icom"
            st.session_state.icom_taxonomy_text = taxonomy_to_text(ICOM_KEYWORDS)
            st.rerun()

law_taxonomy = text_to_taxonomy(law_taxonomy_text)
icom_taxonomy = text_to_taxonomy(icom_taxonomy_text)

if not law_taxonomy:
    st.warning("El marco 1 no tiene etiquetas validas. Usa el formato: etiqueta = palabra1, palabra2")
if not icom_taxonomy:
    st.warning("El marco 2 no tiene etiquetas validas. Usa el formato: etiqueta = palabra1, palabra2")

st.subheader("Cantidad de parrafos representativos")
st.info(
    "La app calcula automaticamente el K optimo con metodo del codo. "
    "Se usa un minimo operativo de 30 y un maximo de 90 parrafos representativos."
)

st.subheader("Clasificacion automatizada con IA")
ollama_model = st.text_input(
    "Modelo Ollama",
    value="mistral:latest",
    help="Modelo local usado para clasificar cada parrafo dentro de la taxonomia cerrada.",
)
ollama_url = st.text_input(
    "URL de Ollama",
    value="http://127.0.0.1:11434",
    help="Endpoint del servidor Ollama. En el remoto normalmente se deja como 127.0.0.1.",
)
st.write(
    "**Entrada de texto** (pega tu texto o sube un .txt/.md). Si hay saltos "
    "de linea, se usan como parrafos; si no, se segmenta por oraciones."
)
col1, col2 = st.columns(2)
with col1:
    texto = st.text_area(
        "Pegar texto",
        height=220,
        placeholder="Pega aqui tu texto en espanol...",
    )
with col2:
    up = st.file_uploader("...o subir archivo .txt/.md", type=["txt", "md"])

if up and not texto.strip():
    texto = up.read().decode("utf-8", errors="replace")

_, process_col, _ = st.columns([1, 2, 1])
with process_col:
    process_clicked = st.button("Procesar", type="primary", use_container_width=True)

if process_clicked:
    if not texto.strip():
        st.error("Por favor, introduce texto o sube un .txt.")
        st.stop()

    items = split_sentences_or_paragraphs(texto)
    if len(items) == 0:
        st.error(
            "No se detectaron parrafos con suficiente texto legible. "
            "Se descartan fragmentos compuestos solo por numeros, codigos o signos."
        )
        st.stop()

    selected_items, elbow_rows, elbow_summary = select_k_medoids_with_elbow(
        items,
        min_k=30,
        max_k=90,
    )
    master_rows = []
    preview_rows = []

    for paragraph in selected_items:
        norm_labels = automated_labeling_columns(
            paragraph,
            law_taxonomy,
            icom_taxonomy,
            ollama_model,
            ollama_url,
        )
        master_rows.append(
            [
                paragraph,
                norm_labels["ley_diccionario"],
                norm_labels["codigo_diccionario"],
                norm_labels["etiqueta_ia_ley_19_2013"],
                norm_labels["etiqueta_ia_codigo_deontologico"],
                norm_labels["justificacion_ia_ley_19_2013"],
                norm_labels["justificacion_ia_codigo_deontologico"],
                norm_labels["razonamiento_ia_ley_19_2013"],
                norm_labels["razonamiento_ia_codigo_deontologico"],
                norm_labels["etiqueta_final"],
            ]
        )
        preview_rows.append(
            {
                "parrafo": paragraph,
                f"{law_name}_diccionario": norm_labels["ley_diccionario"],
                f"{icom_name}_diccionario": norm_labels["codigo_diccionario"],
                "etiqueta_ia_ley_19_2013": norm_labels["etiqueta_ia_ley_19_2013"],
                "etiqueta_ia_codigo_deontologico": norm_labels["etiqueta_ia_codigo_deontologico"],
                "justificacion_ia_ley_19_2013": norm_labels["justificacion_ia_ley_19_2013"],
                "justificacion_ia_codigo_deontologico": norm_labels["justificacion_ia_codigo_deontologico"],
                "razonamiento_ia_ley_19_2013": norm_labels["razonamiento_ia_ley_19_2013"],
                "razonamiento_ia_codigo_deontologico": norm_labels["razonamiento_ia_codigo_deontologico"],
                "etiqueta_final": norm_labels["etiqueta_final"],
            }
        )

    master_excel_bytes = make_excel_bytes(
        master_rows,
        header=[
            "parrafo",
            f"{law_name}_diccionario",
            f"{icom_name}_diccionario",
            "etiqueta_ia_ley_19_2013",
            "etiqueta_ia_codigo_deontologico",
            "justificacion_ia_ley_19_2013",
            "justificacion_ia_codigo_deontologico",
            "razonamiento_ia_ley_19_2013",
            "razonamiento_ia_codigo_deontologico",
            "etiqueta_final",
        ],
        elbow_rows=elbow_rows,
        elbow_summary=elbow_summary,
    )

    excel_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    st.success(
        f"Archivo madre listo con {len(selected_items)} parrafos representativos "
        f"(K optimo por codo = {elbow_summary['k_optimo']})."
    )
    st.download_button(
        "Descargar archivo_madre_etiquetado.xlsx",
        data=master_excel_bytes,
        file_name="archivo_madre_etiquetado.xlsx",
        mime=excel_mime,
    )

    st.subheader("Vista previa")
    st.write("**Archivo madre**")
    st.dataframe(preview_rows[: min(10, len(preview_rows))])
