#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import re
import random
import argparse
import math
import unicodedata
from collections import Counter
from pathlib import Path
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

DEFAULT_MODEL = "dccuchile/bert-base-spanish-wwm-cased"


LEY_19_2013_KEYWORDS = {
    "informacion_institucional_funciones_institucionales": ["funciones institucionales", "funciones", "cometido"],
    "informacion_institucional_fines_objetivos": ["fines", "objetivos institucionales", "fines y objetivos institucionales", "mision", "finalidad"],
    "informacion_institucional_marco_normativo": ["marco normativo", "normativa aplicable", "regimen juridico"],
    "informacion_institucional_naturaleza_juridica": ["naturaleza juridica", "fundacion", "consorcio", "organismo", "entidad publica"],
    "informacion_institucional_identidad_datos_basicos": ["identidad", "datos basicos", "denominacion", "sede", "titularidad"],
    "informacion_organizativa_estructura": ["estructura organizativa", "estructura", "departamentos", "areas"],
    "informacion_organizativa_organigrama": ["organigrama"],
    "informacion_organizativa_responsables_organos": ["responsables", "organos", "identificacion de responsables", "responsables de los distintos organos", "direccion", "gerencia", "jefatura"],
    "informacion_organizativa_perfil_trayectoria_altos_cargos": ["perfil profesional", "trayectoria profesional", "altos cargos", "maximos responsables"],
    "informacion_organizativa_regimen_dedicacion": ["regimen de dedicacion", "dedicacion", "jornada"],
    "informacion_organizativa_retribuciones_altos_cargos": ["retribuciones", "salario", "remuneracion", "altos cargos"],
    "informacion_organizativa_indemnizaciones_cese": ["indemnizacion", "indemnizaciones", "cese"],
    "planificacion_planes_estrategicos": ["plan estrategico", "planes estrategicos", "estrategia"],
    "planificacion_planes_anuales": ["plan anual", "planes anuales"],
    "planificacion_programas_actuacion": ["programa de actuacion", "programas de actuacion"],
    "planificacion_objetivos_anuales": ["objetivos anuales", "objetivo anual"],
    "planificacion_indicadores_asociados": ["indicadores asociados", "indicadores del plan", "indicadores"],
    "planificacion_actividades": ["planificacion de actividades", "programacion", "calendario de actividades"],
    "evaluacion_informes_seguimiento": ["informe de seguimiento", "informes de seguimiento", "seguimiento de planes"],
    "evaluacion_indicadores_ejecucion": ["indicadores de ejecucion", "ejecucion"],
    "evaluacion_resultados_obtenidos": ["resultados obtenidos", "resultados"],
    "evaluacion_nivel_cumplimiento_objetivos": ["nivel de cumplimiento", "cumplimiento de objetivos", "grado de cumplimiento"],
    "evaluacion_interna_externa": ["evaluacion interna", "evaluacion externa", "auditoria", "evaluaciones"],
    "normativa_leyes_aplicables": ["ley aplicable", "leyes aplicables", "ley"],
    "normativa_reglamentos": ["reglamento", "reglamentos"],
    "normativa_estatutos": ["estatuto", "estatutos"],
    "normativa_sectorial_especifica": ["normativa sectorial", "normativa especifica"],
    "normativa_proyectos_en_tramitacion": ["proyecto normativo", "proyectos normativos", "tramitacion", "informacion publica"],
    "funciones_competencias_atribuidas": ["competencias atribuidas", "competencias", "atribuciones"],
    "funciones_servicios_prestados": ["servicios prestados", "servicios", "prestacion"],
    "funciones_actividades_realizadas": ["actividades realizadas", "actividades", "actuaciones"],
    "funciones_ambito_territorial": ["ambito territorial", "territorio", "actuacion territorial"],
    "informacion_juridica_directrices": ["directriz", "directrices"],
    "informacion_juridica_instrucciones": ["instruccion", "instrucciones"],
    "informacion_juridica_circulares": ["circular", "circulares"],
    "informacion_juridica_acuerdos": ["acuerdo", "acuerdos"],
    "informacion_juridica_respuestas_consultas": ["respuesta a consulta", "consultas relevantes"],
    "informacion_juridica_documentos_informacion_publica": ["documentos sometidos a informacion publica", "informacion publica"],
    "informacion_juridica_convenios_suscritos": ["convenio suscrito", "convenios suscritos", "objeto", "partes firmantes", "duracion", "obligaciones economicas"],
    "informacion_juridica_encomiendas_gestion": ["encomienda de gestion", "encomiendas de gestion"],
    "contratos_objeto": ["objeto del contrato", "objeto contractual"],
    "contratos_duracion": ["duracion del contrato", "plazo del contrato"],
    "contratos_importe_licitacion_adjudicacion": ["importe de licitacion", "importe de adjudicacion", "licitacion", "adjudicacion"],
    "contratos_procedimiento_utilizado": ["procedimiento utilizado", "procedimiento de contratacion"],
    "contratos_identidad_adjudicatario": ["adjudicatario", "identidad del adjudicatario"],
    "contratos_modificaciones_contractuales": ["modificacion contractual", "modificaciones contractuales"],
    "convenios_partes_firmantes": ["partes firmantes", "firmantes"],
    "convenios_objeto": ["objeto del convenio", "objeto"],
    "convenios_duracion": ["duracion del convenio", "vigencia"],
    "convenios_obligaciones_economicas": ["obligaciones economicas", "aportacion economica"],
    "subvenciones_importe": ["importe", "cuantia", "subvencion", "subvenciones", "ayudas publicas"],
    "subvenciones_objetivo": ["objetivo de la subvencion", "finalidad de la ayuda"],
    "subvenciones_beneficiarios": ["beneficiario", "beneficiarios"],
    "presupuestos_presupuesto_anual": ["presupuesto anual", "presupuesto"],
    "presupuestos_estado_ejecucion": ["estado de ejecucion presupuestaria", "ejecucion presupuestaria"],
    "cuentas_anuales_balance": ["balance"],
    "cuentas_anuales_cuenta_resultados": ["cuenta de resultados", "resultado economico"],
    "cuentas_anuales_memoria": ["memoria", "memoria economica", "cuentas anuales"],
    "cuentas_anuales_informes_auditoria": ["informe de auditoria", "auditoria"],
    "estadisticas_datos_actividad": ["datos de actividad", "actividad relevante"],
    "estadisticas_indicadores_publicos": ["indicadores publicos", "indicadores"],
    "estadisticas_informacion_agregada": ["informacion agregada", "interes general"],
    "acceso_procedimiento_derecho_acceso": ["derecho de acceso", "procedimiento de acceso", "solicitud de acceso"],
    "acceso_organo_competente": ["organo competente", "organo competente para resolver solicitudes", "resolver solicitudes"],
    "acceso_plazos_respuesta": ["plazo de respuesta", "plazos de respuesta"],
    "acceso_recursos_posibles": ["recurso", "recursos posibles", "reclamacion"],
    "acceso_estadisticas_solicitudes": ["solicitudes recibidas", "solicitudes resueltas", "estadisticas sobre solicitudes"],
    "acceso_portal_transparencia": ["portal de transparencia", "transparencia"],
    "buen_gobierno_principios_actuacion": ["principios de actuacion", "altos cargos"],
    "buen_gobierno_conflicto_intereses": ["conflicto de intereses", "incompatibilidades"],
    "buen_gobierno_regimen_sancionador": ["regimen sancionador", "sancion"],
    "buen_gobierno_declaraciones_bienes_actividades": ["declaracion de bienes", "declaraciones de bienes", "declaracion de actividades"],
    "buen_gobierno_codigo_etico_conducta": ["codigo etico", "codigo de conducta", "conducta"],
}


ICOM_KEYWORDS = {
    "dimension_01_transparencia_institucional": [
        "mision", "vision", "objetivos institucionales", "titularidad",
        "naturaleza juridica", "organos de gobierno", "direccion",
        "organigrama", "plan estrategico", "politicas institucionales",
    ],
    "dimension_02_transparencia_economica": [
        "presupuesto", "fuentes de financiacion", "subvenciones",
        "patrocinios", "patrocinio", "mecenazgo", "recursos economicos",
        "gestion financiera", "sostenibilidad economica",
    ],
    "dimension_03_transparencia_colecciones": [
        "politica de colecciones", "inventarios", "inventario", "catalogo",
        "documentacion", "conservacion preventiva", "restauracion",
        "adquisiciones", "adquisicion", "donaciones", "donacion",
        "depositos", "deposito", "prestamos", "prestamo",
        "bajas de coleccion", "baja de coleccion", "digitalizacion",
        "objetos digitales", "banco de imagenes", "imagenes archivisticas",
        "imagen archivistica",
    ],
    "dimension_04_transparencia_procedencia_legalidad": [
        "procedencia", "titularidad", "autenticidad", "diligencia debida",
        "trafico ilicito", "exportacion", "importacion", "restitucion",
        "repatriacion",
    ],
    "dimension_05_transparencia_cientifica": [
        "proyectos de investigacion", "proyecto de investigacion",
        "publicaciones", "publicacion", "produccion cientifica",
        "bases de datos", "base de datos", "documentacion cientifica",
        "investigadores", "investigador", "innovacion", "difusion cientifica",
    ],
    "dimension_06_transparencia_educativa": [
        "programas educativos", "programa educativo", "talleres", "taller",
        "actividades escolares", "actividad escolar", "visitas guiadas",
        "visita guiada", "recursos didacticos", "recurso didactico",
        "mediacion cultural", "divulgacion", "exposiciones temporales",
        "exposicion temporal", "actividades culturales", "actividad cultural",
    ],
    "dimension_07_transparencia_acceso": [
        "horarios", "horario", "tarifas", "tarifa", "accesibilidad fisica",
        "accesibilidad digital", "colecciones en linea", "coleccion en linea",
        "catalogo digital", "derechos de uso", "servicios al visitante",
        "servicio al visitante",
    ],
    "dimension_08_transparencia_servicio_publico": [
        "biblioteca", "archivo", "centro de documentacion",
        "servicios para investigadores", "servicio para investigadores",
        "cesion de espacios", "actividades para la comunidad",
        "actividad para la comunidad", "servicios especializados",
        "servicio especializado",
    ],
    "dimension_09_transparencia_social": [
        "participacion ciudadana", "comunidad local", "inclusion",
        "diversidad", "accesibilidad social", "voluntariado",
        "responsabilidad social", "participacion de grupos de interes",
    ],
    "dimension_10_transparencia_cooperacion_institucional": [
        "convenios", "convenio", "redes nacionales", "red nacional",
        "redes internacionales", "red internacional", "universidades",
        "universidad", "centros de investigacion", "centro de investigacion",
        "proyectos colaborativos", "proyecto colaborativo",
        "cooperacion institucional", "prestamos interinstitucionales",
        "prestamo interinstitucional",
    ],
    "dimension_11_transparencia_juridica": [
        "cumplimiento normativo", "normativa aplicable",
        "proteccion del patrimonio", "propiedad intelectual",
        "proteccion de datos", "derechos culturales", "cumplimiento legal",
    ],
    "dimension_12_transparencia_etica": [
        "codigo etico", "integridad institucional", "buen gobierno",
        "conflictos de interes", "conflicto de interes",
        "responsabilidad profesional", "politicas de integridad",
        "mecanismos de denuncia", "mecanismo de denuncia",
        "compromiso etico",
    ],
}


# --------------------------
# Utilidades
# --------------------------
def set_seed(seed: int = 42):
    random.seed(seed)


def read_txt_file(path: str) -> str:
    """Lee texto con tolerancia a codificaciones."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except UnicodeDecodeError:
        try:
            with open(path, "r", encoding="latin-1") as f:
                return f.read()
        except UnicodeDecodeError:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()


def split_sentences_or_paragraphs(text: str) -> List[str]:
    """
    Divide el texto en oraciones o párrafos.
    Usa saltos de línea si existen; si no, divide por puntos.
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) > 1:
        return lines
    # si no hay saltos, segmentar por punto
    sentences = [s.strip() for s in re.split(r"[.!?]\s+", text) if len(s.strip()) > 0]
    return sentences


def mean_pooling(last_hidden_state: torch.Tensor, attention_mask: torch.Tensor) -> torch.Tensor:
    mask = attention_mask.unsqueeze(-1).expand(last_hidden_state.size()).float()
    summed = torch.sum(last_hidden_state * mask, dim=1)
    counts = torch.clamp(mask.sum(dim=1), min=1e-9)
    return summed / counts


def get_bert_embeddings(texts: List[str], model_name: str = DEFAULT_MODEL, device: str = None, batch_size: int = 16) -> np.ndarray:
    """Devuelve embeddings promedio para cada oración/párrafo."""
    if device is None:
        device = "cuda" if torch.cuda.is_available() else "cpu"

    tokenizer = AutoTokenizer.from_pretrained(model_name)
    model = AutoModel.from_pretrained(model_name)
    model.to(device)
    model.eval()

    embs = []
    with torch.no_grad():
        for i in range(0, len(texts), batch_size):
            batch = texts[i:i+batch_size]
            enc = tokenizer(batch, padding=True, truncation=True, return_tensors="pt", max_length=256)
            enc = {k: v.to(device) for k, v in enc.items()}
            out = model(**enc)
            pooled = mean_pooling(out.last_hidden_state, enc["attention_mask"])
            embs.append(pooled.detach().cpu().numpy())

    return np.vstack(embs)


def pairwise_distance_matrix(X: np.ndarray, metric: str = "cosine") -> np.ndarray:
    """Devuelve matriz de distancias (por defecto: coseno)."""
    return pairwise_distances(X, metric=metric)


def normalize_for_keywords(text: str) -> str:
    """Normaliza texto para detectar etiquetas por palabras clave."""
    normalized = unicodedata.normalize("NFKD", text)
    without_accents = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return without_accents.lower()


def keyword_labels(text: str, taxonomy: Dict[str, List[str]]) -> List[str]:
    """Devuelve las etiquetas cuya lista de palabras clave aparece en el texto."""
    normalized_text = normalize_for_keywords(text)
    labels = []
    for label, keywords in taxonomy.items():
        for keyword in keywords:
            if normalize_for_keywords(keyword) in normalized_text:
                labels.append(label)
                break
    return labels


def normative_columns(text: str) -> Dict[str, str]:
    """Clasifica una frase representativa en Ley 19/2013, ICOM u otros."""
    ley_labels = keyword_labels(text, LEY_19_2013_KEYWORDS)
    icom_labels = keyword_labels(text, ICOM_KEYWORDS)
    return {
        "ley 19/2013": "; ".join(ley_labels),
        "codigo deontologico": "; ".join(icom_labels),
        "otros": "" if ley_labels or icom_labels else "otros",
    }


def select_k_medoids_paragraphs(paragraphs: List[str], k: int = 24, seed: int = 42) -> List[str]:
    """Selecciona K parrafos representativos con una aproximacion ligera a K-Medoids."""
    if not paragraphs:
        return []
    if k <= 0:
        k = 1
    if len(paragraphs) <= k:
        return paragraphs

    vectors = tfidf_vectors(paragraphs)
    distances = cosine_distance_matrix(vectors)
    medoid_indices = initialize_medoids(distances, k)
    medoid_indices = improve_medoids(distances, medoid_indices, max_iter=5)
    return [paragraphs[idx] for idx in medoid_indices]


def tokenize_for_tfidf(text: str) -> List[str]:
    normalized = normalize_for_keywords(text)
    return re.findall(r"\b[a-z0-9]{3,}\b", normalized)


def tfidf_vectors(texts: List[str]) -> List[Dict[str, float]]:
    tokenized = [tokenize_for_tfidf(text) for text in texts]
    doc_freq = Counter()
    for tokens in tokenized:
        doc_freq.update(set(tokens))

    total_docs = len(texts)
    vectors = []
    for tokens in tokenized:
        counts = Counter(tokens)
        total_terms = sum(counts.values()) or 1
        vector = {}
        for token, count in counts.items():
            tf = count / total_terms
            idf = math.log((1 + total_docs) / (1 + doc_freq[token])) + 1
            vector[token] = tf * idf
        norm = math.sqrt(sum(value * value for value in vector.values())) or 1
        vectors.append({token: value / norm for token, value in vector.items()})
    return vectors


def cosine_distance(vec_a: Dict[str, float], vec_b: Dict[str, float]) -> float:
    if len(vec_a) > len(vec_b):
        vec_a, vec_b = vec_b, vec_a
    similarity = sum(value * vec_b.get(token, 0.0) for token, value in vec_a.items())
    return 1.0 - similarity


def cosine_distance_matrix(vectors: List[Dict[str, float]]) -> List[List[float]]:
    size = len(vectors)
    distances = [[0.0] * size for _ in range(size)]
    for i in range(size):
        for j in range(i + 1, size):
            distance = cosine_distance(vectors[i], vectors[j])
            distances[i][j] = distance
            distances[j][i] = distance
    return distances


def initialize_medoids(distances: List[List[float]], k: int) -> List[int]:
    total_distances = [sum(row) for row in distances]
    medoids = [min(range(len(distances)), key=lambda idx: total_distances[idx])]
    while len(medoids) < k:
        candidates = [idx for idx in range(len(distances)) if idx not in medoids]
        next_medoid = max(candidates, key=lambda idx: min(distances[idx][m] for m in medoids))
        medoids.append(next_medoid)
    return medoids


def medoid_cost(distances: List[List[float]], medoids: List[int]) -> float:
    return sum(min(row[medoid] for medoid in medoids) for row in distances)


def improve_medoids(distances: List[List[float]], medoids: List[int], max_iter: int = 5) -> List[int]:
    medoids = list(medoids)
    best_cost = medoid_cost(distances, medoids)
    all_indices = set(range(len(distances)))

    for _ in range(max_iter):
        improved = False
        non_medoids = sorted(all_indices.difference(medoids))
        for medoid in list(medoids):
            medoid_position = medoids.index(medoid)
            for candidate in non_medoids:
                trial = list(medoids)
                trial[medoid_position] = candidate
                cost = medoid_cost(distances, trial)
                if cost < best_cost:
                    medoids = trial
                    best_cost = cost
                    improved = True
                    break
            if improved:
                break
        if not improved:
            break

    return sorted(medoids)


def write_xlsx(path: str, header: List[str], rows: List[List]):
    """Escribe una hoja Excel simple con cabecera y filas."""
    wb = Workbook()
    ws = wb.active
    ws.title = "archivo_madre"
    ws.append(header)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25

    wb.save(path)


def elbow_curve(D: np.ndarray, max_k: int = 12, seed: int = 42) -> Dict[int, float]:
    """Calcula curva elbow para distintos K."""
    max_k = min(max_k, D.shape[0])
    results = {}
    for k in range(1, max_k + 1):
        model = KMedoids(n_clusters=k, metric="precomputed", random_state=seed)
        model.fit(D)
        labels_k = model.labels_
        medoids_k = model.medoid_indices_
        cost = float(sum(D[i, medoids_k[labels_k[i]]] for i in range(D.shape[0])))
        results[k] = cost
    return results


# --------------------------
# MAIN
# --------------------------
def main():
    parser = argparse.ArgumentParser(description="Seleccion de parrafos representativos con K-Medoids y etiquetado normativo")
    parser.add_argument("--file", type=str, required=True, help="Ruta al archivo .txt")
    parser.add_argument("--k", type=int, default=24, help="Numero de parrafos representativos que quieres obtener")
    parser.add_argument("--max_k", type=int, default=35, help="Máximo K para el elbow")
    parser.add_argument("--model", type=str, default=DEFAULT_MODEL, help="Modelo BERT de Transformers")
    parser.add_argument("--seed", type=int, default=42, help="Semilla aleatoria")
    parser.add_argument("--save_dir", type=str, default="./resultados_kmedoids", help="Carpeta de salida")
    parser.add_argument("--distance_metric", type=str, default="cosine", choices=["cosine", "euclidean"], help="Métrica de distancia")
    args = parser.parse_args()

    set_seed(args.seed)
    os.makedirs(args.save_dir, exist_ok=True)

    base_name = Path(args.file).stem
    print(f"\nProcesando archivo: {args.file}")
    print(f"Guardando en: {args.save_dir}/")

    # --- Lectura y segmentación ---
    text = read_txt_file(args.file)
    items = split_sentences_or_paragraphs(text)
    selected_items = select_k_medoids_paragraphs(items, k=args.k, seed=args.seed)
    print(f"Total de oraciones/parrafos detectados: {len(items)}")
    print(f"Parrafos representativos seleccionados (K): {len(selected_items)}")

    # --- Archivo madre: parrafos completos + etiquetas normativas ---
    master_xlsx = os.path.join(args.save_dir, f"{base_name}_archivo_madre_etiquetado.xlsx")
    master_rows = []
    for paragraph in selected_items:
        norm_labels = normative_columns(paragraph)
        master_rows.append([
            paragraph,
            norm_labels["ley 19/2013"],
            norm_labels["codigo deontologico"],
            norm_labels["otros"],
        ])
    write_xlsx(
        master_xlsx,
        ["parrafo", "ley 19/2013", "codigo deontologico icom", "otros temas"],
        master_rows,
    )
    print(f"[OK] Guardado archivo madre: {master_xlsx}")

    print(f"\nCompletado: {base_name}")
    print("Archivo generado:")
    print(f"   - {os.path.basename(master_xlsx)}")


if __name__ == "__main__":
    main()
